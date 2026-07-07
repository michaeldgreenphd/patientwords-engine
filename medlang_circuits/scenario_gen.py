"""Claude-driven scenario generation + the hand-built spreadsheet importer.

Four capabilities behind the ``medlang-generate`` CLI:

- ``generate_stress_pairs`` - have Claude author new patient-vs-clinical
  next-token stress pairs in the hand-built dataset's format (single
  contiguous term swap inside an identical syntactic frame, ending at a
  next-token probe boundary), programmatically validated before acceptance.
- ``generate_quadrant_scenarios`` - have Claude author 4-quadrant items for
  the generic morphosyntax-x-lexicon matrix: a standard/nonstandard frame
  pair with a single {term} slot, crossed with a medical/patient term pair
  (feeds ``--mode 4quadrant`` in medlang-batch-eval).
- ``generate_dialect_variants`` - hold one swapped term verbatim and have
  Claude rewrite only the surrounding syntax as different English
  dialects/registers would frame the same situation (feeds ``--mode dialect``
  in medlang-batch-eval).
- ``import_sheet`` - convert the hand-built spreadsheet (Phrase 1 | Next
  token | Prob | Link | Phrase 2 | ... | Notes | sourcing) into batch-ready
  pair JSON, carrying the manual measurements along as provenance.

Repo rules honored here: model selection via argument or
MEDLANG_ANTHROPIC_MODEL; hard spend ceiling via evaluate_models.CostTracker;
the Messages call goes through the evaluate_models._call seam so tests run
offline; and NO medical vocabulary lives in this source file - all concrete
phrases arrive from the LLM, seed files, or the spreadsheet, and leave as
JSON data files.
"""

from __future__ import annotations

import argparse
import csv
import difflib
import json
import logging
import math
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator

from medlang_circuits.evaluate_models import CostTracker, _call, resolve_models
from medlang_circuits.llm_client import DEFAULT_MODEL, _get_client

logger = logging.getLogger(__name__)

TERMINAL_PUNCTUATION = ".!?…;:"
GEN_BATCH_SIZE = 6  # candidates requested per API call
# Sized so a full batch of candidates with rationales never hits the cap: a
# 1600-token cap truncated every round of a 6-candidate run to unparseable
# JSON (0 accepted from 6 rounds).
GEN_MAX_TOKENS = 4096
DIALECT_MAX_TOKENS = 2048

# Register/dialect coverage when the caller doesn't pass --dialects. These are
# framing varieties, not caricatures - the prompt enforces respectful renderings.
DEFAULT_DIALECTS = (
    "Southern US English",
    "African American Vernacular English",
    "New England US English",
    "British English",
    "Irish English",
    "Caribbean English",
    "ESL-influenced English",
    "texting/informal register",
    "terse spoken register",
    "elaborated formal register",
)

STRESS_PAIR_SYSTEM = (
    "You generate stress-test items for a mechanistic-interpretability study of how a small "
    "language model handles colloquial patient language versus precise clinical terminology. "
    "Rules for every item - follow all of them exactly:\n"
    "1. An item is TWO phrases sharing an IDENTICAL syntactic frame; only one contiguous term "
    "span differs between them. Phrase 1 uses a colloquial/patient expression (an idiom, slang, "
    "or lay description of a symptom, condition, or device); phrase 2 replaces that span with "
    "its precise clinical equivalent, in the same position. Do not change anything else.\n"
    "2. Both phrases end mid-sentence at a natural next-token probe point - typically right "
    "after an article or possessive (for example '...so I need to talk to a', '...I need to "
    "take my', '...so I should take some'). No trailing period or any terminal punctuation. "
    "The next token is left blank: it is the measurement, not part of the item.\n"
    "3. The frame must make the next token diagnostic: the natural continuation should "
    "plausibly differ between a mundane everyday reading and a care-seeking/medical reading "
    "of the situation, so the term swap can move which continuation the model prefers.\n"
    "4. Vary the frames across items: first and third person, different tenses, different "
    "probe words, everyday settings (work, a bar, the weather, family life).\n"
    "5. The FIRST expected_clinical_continuations entry is the measured target: it must be a "
    "single, common, high-frequency English word (ideally one wordpiece - e.g. an everyday noun "
    "right after an article or possessive), the kind of continuation a small 2B-parameter model "
    "would place in its top-10 next tokens for the clinical phrasing. No rare drug names, no "
    "multi-word targets, no proper nouns. Design the frame so its grammar funnels hard toward "
    "that word (for example 'taking an' before a vowel-initial remedy).\n"
    "Output STRICT JSON only - a JSON array of objects, each with exactly these keys: "
    '{"patient_prompt": str, "clinical_prompt": str, "patient_term": str, "clinical_term": str, '
    '"expected_clinical_continuations": [str, ...], "topic": str, "rationale": str} - '
    '"topic" names the single condition area this item covers (one of the steered topics '
    "when topics are given). No text outside the JSON."
)

QUADRANT_SYSTEM = (
    "You generate four-way stress-test items for a mechanistic-interpretability study of how a "
    "small language model treats medical lexicon versus patient-derived language, crossed with "
    "standard versus nonstandard English morphosyntax. Each item is a FRAME PAIR crossed with a "
    "TERM PAIR - follow all rules exactly:\n"
    "1. 'standard_frame' and 'nonstandard_frame' describe the SAME situation and contain the "
    "placeholder {term} exactly once, at the same semantic position. The standard frame uses "
    "standard English morphosyntax. The nonstandard frame re-renders ONLY the grammar (verbal "
    "aspect, agreement, copula, tense marking) the way a real speaker of a nonstandard English "
    "variety would - authentic and respectful, never caricature, exaggeration, or mockery. Keep "
    "the content words identical apart from the grammar.\n"
    "2. 'medical_term' is the precise clinical lexicon for the condition or situation; "
    "'patient_term' is the colloquial patient-derived expression for the same thing. Each must "
    "fit the {term} slot of BOTH frames grammatically.\n"
    "3. Both frames end mid-sentence at a natural next-token probe point (for example after "
    "'...I should go to', '...I need to take my', '...she picked up some'), with no terminal "
    "punctuation. The next token is the measurement, not part of the item.\n"
    "4. The frame must make the next token diagnostic: the natural continuation should plausibly "
    "differ between a mundane everyday reading and a care-seeking/medical reading.\n"
    "5. Vary persons, tenses, probe words, and everyday settings across items.\n"
    "Output STRICT JSON only - a JSON array of objects, each with exactly these keys: "
    '{"standard_frame": str, "nonstandard_frame": str, "medical_term": str, "patient_term": str, '
    '"expected_clinical_continuations": [str, ...], "rationale": str}. Abstract shape example '
    '(structure only, invent real content): {"standard_frame": "I have{term}, and I feel really '
    'unwell. I should go to", "nonstandard_frame": "I been had{term}, and I been feeling really '
    'unwell. I should go to", "medical_term": " alpha", "patient_term": " the beta"}. '
    "No text outside the JSON."
)

DIALECT_SYSTEM = (
    "You rewrite one phrase into dialect and register variants for a mechanistic-"
    "interpretability study of how a language model treats different framings of the same "
    "clinical situation. Requirements - follow all of them exactly:\n"
    "1. Renderings must be authentic and respectful: no caricature, no exaggerated "
    "stereotype, no mockery. Each variant simulates how a real patient who speaks that "
    "variety would naturally phrase the situation.\n"
    "2. The fixed term you are given must appear UNCHANGED, verbatim, in every variant. "
    "Rewrite ONLY the surrounding syntax and framing.\n"
    "3. Every variant must end mid-sentence at a next-token probe boundary comparable to "
    "the original (e.g. after the same kind of article or possessive), with no terminal "
    "punctuation, so the next token stays measurable.\n"
    "4. Keep the described situation and its meaning identical; only the dialect/register "
    "of the framing changes.\n"
    "Output STRICT JSON only - a JSON array of objects, each with exactly these keys: "
    '{"dialect": str, "prompt": str}. No text outside the JSON.'
)


# ---------------------------------------------------------------------------
# Shared validation helpers (pure functions, unit-tested offline)
# ---------------------------------------------------------------------------


def ends_at_probe_boundary(text: str | None) -> bool:
    """True if the phrase ends mid-sentence (non-empty, no terminal punctuation)."""
    stripped = (text or "").strip()
    return bool(stripped) and stripped[-1] not in TERMINAL_PUNCTUATION


def single_span_swap(patient_prompt: str, clinical_prompt: str) -> dict[str, str] | None:
    """Token-level diff of the two prompts; the swapped spans if it is a single
    contiguous replacement (the single-swap property of the hand-built set), else None."""
    a, b = patient_prompt.split(), clinical_prompt.split()
    matcher = difflib.SequenceMatcher(a=a, b=b, autojunk=False)
    edits = [op for op in matcher.get_opcodes() if op[0] != "equal"]
    if len(edits) != 1 or edits[0][0] != "replace":
        return None
    _, i1, i2, j1, j2 = edits[0]
    return {"patient_span": " ".join(a[i1:i2]), "clinical_span": " ".join(b[j1:j2])}


def _norm_prompt(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().casefold())


def _pair_key(patient_prompt: str, clinical_prompt: str) -> tuple[str, str]:
    return _norm_prompt(patient_prompt), _norm_prompt(clinical_prompt)


REQUIRED_PAIR_KEYS = ("patient_prompt", "clinical_prompt", "patient_term", "clinical_term",
                      "expected_clinical_continuations")


def validate_stress_pair(candidate: Any, seen: set[tuple[str, str]]) -> tuple[dict[str, Any] | None, str | None]:
    """Programmatic acceptance gate for one generated candidate.

    Returns (batch_ready_pair, None) on acceptance - and adds its dedupe key to
    ``seen`` - or (None, rejection_reason) otherwise.
    """
    if not isinstance(candidate, dict):
        return None, "not an object"
    missing = [k for k in REQUIRED_PAIR_KEYS if not candidate.get(k)]
    if missing:
        return None, f"missing keys: {missing}"
    continuations = candidate["expected_clinical_continuations"]
    if not isinstance(continuations, list) or not any(str(c).strip() for c in continuations):
        return None, "expected_clinical_continuations must be a non-empty list"
    patient_prompt = str(candidate["patient_prompt"]).strip()
    clinical_prompt = str(candidate["clinical_prompt"]).strip()
    for label, prompt in (("patient_prompt", patient_prompt), ("clinical_prompt", clinical_prompt)):
        if not ends_at_probe_boundary(prompt):
            return None, f"{label} does not end at a probe boundary (terminal punctuation or empty)"
    spans = single_span_swap(patient_prompt, clinical_prompt)
    if spans is None:
        return None, "prompts do not differ by a single contiguous term span"
    for label, term, prompt in (("patient_term", candidate["patient_term"], patient_prompt),
                                ("clinical_term", candidate["clinical_term"], clinical_prompt)):
        if str(term).strip().casefold() not in prompt.casefold():
            return None, f"{label} does not appear in its prompt"
    key = _pair_key(patient_prompt, clinical_prompt)
    if key in seen:
        return None, "duplicate of a seed or an already-accepted pair"
    seen.add(key)

    first = next(str(c).strip() for c in continuations if str(c).strip())
    return {
        # batch-ready 2panel schema: clinical wording on top, patient below
        "top_prompt": clinical_prompt,
        "bottom_prompt": patient_prompt,
        "target_clinical_token": " " + first,
        "generation": {
            "patient_term": str(candidate["patient_term"]).strip(),
            "clinical_term": str(candidate["clinical_term"]).strip(),
            "expected_clinical_continuations": [str(c).strip() for c in continuations if str(c).strip()],
            "rationale": str(candidate.get("rationale") or "").strip(),
            "topic": str(candidate.get("topic") or "").strip() or None,
            "swap_spans": spans,
        },
    }, None


REQUIRED_QUADRANT_KEYS = ("standard_frame", "nonstandard_frame", "medical_term", "patient_term",
                          "expected_clinical_continuations")


def _compose_quadrants(frames: dict[str, Any], terms: dict[str, Any]) -> dict[str, str] | None:
    """Compose the four cell prompts (legacy clinical/patient key aliases accepted)."""
    standard = frames.get("standard") or frames.get("clinical")
    nonstandard = frames.get("nonstandard") or frames.get("patient")
    medical = terms.get("medical") or terms.get("clinical")
    patient = terms.get("patient")
    if not all((standard, nonstandard, medical, patient)):
        return None
    return {
        "A": standard.replace("{term}", medical),
        "B": nonstandard.replace("{term}", medical),
        "C": standard.replace("{term}", patient),
        "D": nonstandard.replace("{term}", patient),
    }


def validate_quadrant_item(candidate: Any, seen: set[tuple[str, str]]) -> tuple[dict[str, Any] | None, str | None]:
    """Acceptance gate for one generated 4-quadrant item (frames x terms).

    Returns (batch_ready_item, None) on acceptance - adding its dedupe key to
    ``seen`` - or (None, rejection_reason).
    """
    if not isinstance(candidate, dict):
        return None, "not an object"
    missing = [k for k in REQUIRED_QUADRANT_KEYS if not candidate.get(k)]
    if missing:
        return None, f"missing keys: {missing}"
    continuations = candidate["expected_clinical_continuations"]
    if not isinstance(continuations, list) or not any(str(c).strip() for c in continuations):
        return None, "expected_clinical_continuations must be a non-empty list"

    frames = {}
    for key in ("standard_frame", "nonstandard_frame"):
        frame = re.sub(r"[ \t]+\{term\}", "{term}", str(candidate[key]).strip())
        if frame.count("{term}") != 1:
            return None, f"{key} must contain the placeholder {{term}} exactly once"
        if not ends_at_probe_boundary(frame):
            return None, f"{key} does not end at a probe boundary (terminal punctuation or empty)"
        frames[key] = frame
    if _norm_prompt(frames["standard_frame"]) == _norm_prompt(frames["nonstandard_frame"]):
        return None, "frames are identical - no morphosyntax shift"

    terms = {}
    for key in ("medical_term", "patient_term"):
        term = str(candidate[key]).strip()
        if not term:
            return None, f"{key} is empty"
        terms[key] = " " + term  # composition relies on the term carrying its leading space
    if terms["medical_term"].casefold() == terms["patient_term"].casefold():
        return None, "terms are identical - no lexicon shift"

    composed = _compose_quadrants(
        {"standard": frames["standard_frame"], "nonstandard": frames["nonstandard_frame"]},
        {"medical": terms["medical_term"], "patient": terms["patient_term"]},
    )
    key = (_norm_prompt(composed["A"]), _norm_prompt(composed["D"]))
    if key in seen:
        return None, "duplicate of a seed or an already-accepted item"
    seen.add(key)

    first = next(str(c).strip() for c in continuations if str(c).strip())
    return {
        # batch-ready 4quadrant schema: frames x terms compose the four cells
        "frames": {"standard": frames["standard_frame"], "nonstandard": frames["nonstandard_frame"]},
        "terms": {"medical": terms["medical_term"], "patient": terms["patient_term"]},
        "target_clinical_token": " " + first,
        "generation": {
            "medical_term": terms["medical_term"].strip(),
            "patient_term": terms["patient_term"].strip(),
            "expected_clinical_continuations": [str(c).strip() for c in continuations if str(c).strip()],
            "rationale": str(candidate.get("rationale") or "").strip(),
            "quadrants": composed,
        },
    }, None


def validate_dialect_variant(candidate: Any, term: str, seen: set[str]) -> tuple[dict[str, str] | None, str | None]:
    """Acceptance gate for one dialect variant: labeled, term verbatim, probe boundary, new."""
    if not isinstance(candidate, dict):
        return None, "not an object"
    dialect = str(candidate.get("dialect") or "").strip()
    prompt = str(candidate.get("prompt") or "").strip()
    if not dialect:
        return None, "missing dialect label"
    if not prompt:
        return None, "missing prompt"
    if term not in prompt:
        return None, "fixed term is not present verbatim"
    if not ends_at_probe_boundary(prompt):
        return None, "prompt does not end at a probe boundary (terminal punctuation)"
    key = _norm_prompt(prompt)
    if key in seen:
        return None, "duplicate prompt"
    seen.add(key)
    return {"dialect": dialect, "prompt": prompt}, None


def _salvage_truncated_array(segment: str) -> list[Any]:
    """Recover the complete leading objects of an array cut off mid-item (a
    max_tokens truncation): close the array at each '}' from the end until it
    parses. The lost tail is at most one candidate; the rest of the round's
    output stays usable."""
    for cut in range(len(segment) - 1, -1, -1):
        if segment[cut] != "}":
            continue
        try:
            parsed = json.loads(segment[:cut + 1] + "]")
        except json.JSONDecodeError:
            continue
        return parsed if isinstance(parsed, list) else []
    return []


def _parse_json_array(text: str) -> list[Any]:
    """Parse the model's response as a JSON array, tolerating code fences,
    surrounding prose, and mid-item truncation."""
    stripped = re.sub(r"^\s*```(?:json)?\s*|\s*```\s*$", "", text.strip(), flags=re.S)
    try:
        parsed = json.loads(stripped)
    except json.JSONDecodeError:
        start, end = stripped.find("["), stripped.rfind("]")
        if start == -1:
            return []
        if end > start:
            try:
                parsed = json.loads(stripped[start:end + 1])
            except json.JSONDecodeError:
                parsed = _salvage_truncated_array(stripped[start:])
        else:
            parsed = _salvage_truncated_array(stripped[start:])
    return parsed if isinstance(parsed, list) else []


def _require_client(client: Any) -> Any:
    if client is None:
        client = _get_client()
    if client is None:
        raise RuntimeError("Anthropic client unavailable - set ANTHROPIC_API_KEY and install the anthropic package")
    return client


def _resolve_model(model: str | None) -> str:
    """Argument -> MEDLANG_ANTHROPIC_MODEL -> default, with legacy names remapped."""
    resolved, _ = resolve_models([model or DEFAULT_MODEL])
    return resolved[0]


# ---------------------------------------------------------------------------
# 1. Stress-pair generation
# ---------------------------------------------------------------------------


def _seed_examples(seed_pairs: Iterable[dict[str, Any]] | None) -> list[dict[str, Any]]:
    """Few-shot examples from seeds, accepting either the generator or batch schema."""
    examples = []
    for seed in seed_pairs or []:
        patient = seed.get("patient_prompt") or seed.get("bottom_prompt")
        clinical = seed.get("clinical_prompt") or seed.get("top_prompt")
        if not patient or not clinical:
            continue
        example = {"patient_prompt": patient, "clinical_prompt": clinical}
        for key in ("patient_term", "clinical_term", "expected_clinical_continuations", "rationale"):
            value = seed.get(key) or (seed.get("generation") or {}).get(key)
            if value:
                example[key] = value
        target = seed.get("target_clinical_token")
        if target and "expected_clinical_continuations" not in example:
            example["expected_clinical_continuations"] = [target.strip()]
        examples.append(example)
    return examples


def generate_stress_pairs(
    n: int,
    model: str | None = None,
    seed_pairs: list[dict[str, Any]] | None = None,
    topics: list[str] | None = None,
    max_spend: float = 2.0,
    client: Any = None,
    feedback: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Generate up to ``n`` validated stress pairs; returns pairs + rejects + usage.

    Accepted pairs are batch-ready 2panel items (top_prompt / bottom_prompt /
    target_clinical_token from the first expected continuation) so the output
    file feeds medlang-batch-eval directly. ``feedback`` closes the evidence
    loop: entries describing frames the traced model screened out (each with
    the clinical prompt, the intended target, and the model's actual top
    continuations) are shown as counterexamples so the next round designs
    around the model's real behavior instead of repeating the failure.
    """
    model = _resolve_model(model)
    client = _require_client(client)
    tracker = CostTracker(max_spend=max_spend)
    seen: set[tuple[str, str]] = set()
    examples = _seed_examples(seed_pairs)
    for example in examples:
        seen.add(_pair_key(example["patient_prompt"], example["clinical_prompt"]))

    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    rounds = 0
    max_rounds = max(4, 2 * math.ceil(n / GEN_BATCH_SIZE) + 2)
    while len(accepted) < n and rounds < max_rounds and tracker.can_afford(model):
        rounds += 1
        want = min(n - len(accepted), GEN_BATCH_SIZE)
        parts = [f"Generate {want} new items."]
        if topics:
            parts.append("Steer coverage across these topics/situations: " + ", ".join(topics) + ".")
        if feedback:
            parts.append(
                "MEASUREMENT-SCREENING FAILURES from a previous round: for these frames the "
                "traced model's top continuations stayed mundane even under the CLINICAL "
                "wording, so the intended target was unmeasurable. Study why (the probe word "
                "invited a non-medical continuation) and design structurally different frames - "
                "do not imitate or lightly rephrase these:\n" + json.dumps(feedback, indent=2))
        if examples:
            parts.append("Format and quality examples from the existing dataset (do NOT repeat or "
                         "trivially rephrase them):\n" + json.dumps(examples, indent=2))
        used = sorted(key[0] for key in seen)
        if used:
            parts.append("Patient phrasings already used - avoid duplicating any of these:\n"
                         + json.dumps(used, indent=2))
        text, in_tok, out_tok = _call(client, model, STRESS_PAIR_SYSTEM, "\n\n".join(parts),
                                      max_tokens=GEN_MAX_TOKENS)
        tracker.record(model, in_tok, out_tok)
        candidates = _parse_json_array(text)
        if not candidates:
            rejected.append({"candidate": text[:500], "reason": "response was not a JSON array"})
            continue
        for candidate in candidates:
            if len(accepted) >= n:
                break
            pair, reason = validate_stress_pair(candidate, seen)
            if pair is None:
                rejected.append({"candidate": candidate, "reason": reason})
                logger.info("Rejected candidate: %s", reason)
            else:
                pair["generation"]["model"] = model
                if topics:
                    # condition/topic steering rides along so archives can be
                    # categorized later (e.g. per-condition breakdowns)
                    pair["generation"]["topics"] = list(topics)
                accepted.append(pair)
    return {
        "pairs": accepted,
        "topics": list(topics) if topics else [],
        "rejected": rejected,
        "model": model,
        "rounds": rounds,
        "truncated": tracker.truncated,
        "usage": {"total_cost_usd": round(tracker.spent, 6), "per_model": tracker.per_model},
    }


# ---------------------------------------------------------------------------
# 2. Quadrant scenarios (morphosyntax x lexicon; --mode 4quadrant input)
# ---------------------------------------------------------------------------


def _quadrant_seed_examples(seed_items: Iterable[dict[str, Any]] | None) -> list[dict[str, Any]]:
    """Few-shot examples + dedupe keys from existing 4quadrant items."""
    examples = []
    for seed in seed_items or []:
        frames, terms = seed.get("frames") or {}, seed.get("terms") or {}
        composed = _compose_quadrants(frames, terms) if frames and terms else None
        if composed is None:
            continue  # not a 4quadrant item (e.g. a 2panel seed file) - skip
        example = {
            "standard_frame": frames.get("standard") or frames.get("clinical"),
            "nonstandard_frame": frames.get("nonstandard") or frames.get("patient"),
            "medical_term": terms.get("medical") or terms.get("clinical"),
            "patient_term": terms.get("patient"),
        }
        target = seed.get("target_clinical_token")
        continuations = (seed.get("generation") or {}).get("expected_clinical_continuations")
        if continuations:
            example["expected_clinical_continuations"] = continuations
        elif target:
            example["expected_clinical_continuations"] = [target.strip()]
        examples.append((example, composed))
    return examples


def generate_quadrant_scenarios(
    n: int,
    model: str | None = None,
    seed_pairs: list[dict[str, Any]] | None = None,
    topics: list[str] | None = None,
    max_spend: float = 2.0,
    client: Any = None,
) -> dict[str, Any]:
    """Generate up to ``n`` validated 4-quadrant items (frames x terms).

    Accepted items are batch-ready --mode 4quadrant inputs: a standard/
    nonstandard morphosyntax frame pair with a single {term} slot, crossed
    with a medical/patient term pair, plus the target token from the first
    expected continuation.
    """
    model = _resolve_model(model)
    client = _require_client(client)
    tracker = CostTracker(max_spend=max_spend)
    seen: set[tuple[str, str]] = set()
    examples = []
    for example, composed in _quadrant_seed_examples(seed_pairs):
        examples.append(example)
        seen.add((_norm_prompt(composed["A"]), _norm_prompt(composed["D"])))

    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    rounds = 0
    max_rounds = max(4, 2 * math.ceil(n / GEN_BATCH_SIZE) + 2)
    while len(accepted) < n and rounds < max_rounds and tracker.can_afford(model):
        rounds += 1
        want = min(n - len(accepted), GEN_BATCH_SIZE)
        parts = [f"Generate {want} new items."]
        if topics:
            parts.append("Steer coverage across these topics/situations: " + ", ".join(topics) + ".")
        if examples:
            parts.append("Format and quality examples from the existing dataset (do NOT repeat or "
                         "trivially rephrase them):\n" + json.dumps(examples, indent=2))
        used = sorted(key[0] for key in seen)
        if used:
            parts.append("Prestige-form sentences already used - avoid duplicating any of these:\n"
                         + json.dumps(used, indent=2))
        text, in_tok, out_tok = _call(client, model, QUADRANT_SYSTEM, "\n\n".join(parts),
                                      max_tokens=GEN_MAX_TOKENS)
        tracker.record(model, in_tok, out_tok)
        candidates = _parse_json_array(text)
        if not candidates:
            rejected.append({"candidate": text[:500], "reason": "response was not a JSON array"})
            continue
        for candidate in candidates:
            if len(accepted) >= n:
                break
            item, reason = validate_quadrant_item(candidate, seen)
            if item is None:
                rejected.append({"candidate": candidate, "reason": reason})
                logger.info("Rejected quadrant item: %s", reason)
            else:
                item["generation"]["model"] = model
                if topics:
                    item["generation"]["topics"] = list(topics)
                accepted.append(item)
    return {
        "pairs": accepted,
        "topics": list(topics) if topics else [],
        "rejected": rejected,
        "model": model,
        "rounds": rounds,
        "truncated": tracker.truncated,
        "usage": {"total_cost_usd": round(tracker.spent, 6), "per_model": tracker.per_model},
    }


# ---------------------------------------------------------------------------
# 3. Dialect syntax variants
# ---------------------------------------------------------------------------


def generate_dialect_variants(
    phrase: str,
    term: str,
    n_variants: int,
    dialects: list[str] | None = None,
    model: str | None = None,
    held_fixed: str = "clinical",
    max_spend: float = 1.0,
    client: Any = None,
) -> dict[str, Any]:
    """Rewrite ``phrase`` into dialect/register variants with ``term`` held verbatim.

    ``held_fixed`` names which side of the swap the fixed term is: "clinical"
    (the priority direction - isolates the pure syntax/dialect effect) or
    "patient" (compound effect). It is prompt context and output metadata; the
    validation is identical either way.
    """
    if held_fixed not in ("clinical", "patient"):
        raise ValueError(f"held_fixed must be 'clinical' or 'patient'; got {held_fixed!r}")
    if term not in phrase:
        raise ValueError("the fixed term must appear verbatim in the baseline phrase")
    model = _resolve_model(model)
    client = _require_client(client)
    tracker = CostTracker(max_spend=max_spend)
    requested = list(dialects) if dialects else list(DEFAULT_DIALECTS)

    seen: set[str] = {_norm_prompt(phrase)}  # a variant identical to the baseline is useless
    accepted: list[dict[str, str]] = []
    rejected: list[dict[str, Any]] = []
    rounds = 0
    while len(accepted) < n_variants and rounds < 3 and tracker.can_afford(model):
        rounds += 1
        want = n_variants - len(accepted)
        prompt = (
            f"Baseline phrase: {phrase!r}\n"
            f"Fixed term (must appear verbatim, unchanged): {term!r}\n"
            f"The fixed term is the {held_fixed} wording of the swapped span - do not translate "
            "or alter it.\n"
            f"Produce {want} variants, drawing dialects/registers from this list (at most one "
            f"per variety until all are used): {', '.join(requested)}."
        )
        if accepted:
            prompt += "\nAlready produced - do not duplicate:\n" + json.dumps(
                [v["prompt"] for v in accepted], indent=2)
        text, in_tok, out_tok = _call(client, model, DIALECT_SYSTEM, prompt, max_tokens=DIALECT_MAX_TOKENS)
        tracker.record(model, in_tok, out_tok)
        candidates = _parse_json_array(text)
        if not candidates:
            rejected.append({"candidate": text[:500], "reason": "response was not a JSON array"})
            continue
        for candidate in candidates:
            if len(accepted) >= n_variants:
                break
            variant, reason = validate_dialect_variant(candidate, term, seen)
            if variant is None:
                rejected.append({"candidate": candidate, "reason": reason})
                logger.info("Rejected variant: %s", reason)
            else:
                accepted.append(variant)
    return {
        "baseline_prompt": phrase,
        "term": term,
        "held_fixed": held_fixed,
        "variants": accepted,
        "rejected": rejected,
        "model": model,
        "rounds": rounds,
        "truncated": tracker.truncated,
        "usage": {"total_cost_usd": round(tracker.spent, 6), "per_model": tracker.per_model},
    }


def dialect_batch_item(result: dict[str, Any], target_token: str | None = None) -> dict[str, Any]:
    """Shape a generate_dialect_variants result into one --mode dialect batch item."""
    item: dict[str, Any] = {
        "baseline_prompt": result["baseline_prompt"],
        "variants": result["variants"],
        "held_fixed": result["held_fixed"],
        "term": result["term"],
    }
    if target_token:
        item["target_clinical_token"] = target_token
    return item

def _swap_span(a: str, b: str) -> str:
    """The contiguous span of ``a`` that differs from ``b`` (prefix/suffix diff).

    The raw character diff can land mid-word when the swapped words share
    letters ("beta"/"delta" share the "ta" suffix), so the span snaps outward
    to word boundaries before trimming punctuation.
    """
    i = 0
    while i < min(len(a), len(b)) and a[i] == b[i]:
        i += 1
    j = 0
    while j < min(len(a), len(b)) - i and a[len(a) - 1 - j] == b[len(b) - 1 - j]:
        j += 1
    lo, hi = i, len(a) - j
    if lo >= hi:
        return ""
    while lo > 0 and a[lo - 1] not in " \t":
        lo -= 1
    while hi < len(a) and a[hi] not in " \t,.;:!?":
        hi += 1
    return a[lo:hi].strip(" \t,.;:!?")


def generate_dialect_sweep(
    seed_pairs: list[dict[str, Any]],
    num_baselines: int,
    n_variants: int,
    dialects: list[str] | None = None,
    model: str | None = None,
    max_spend: float = 2.0,
    client: Any = None,
) -> dict[str, Any]:
    """Run generate_dialect_variants over several measured baselines.

    Each seed pair contributes its clinical phrasing (``top_prompt``) as the
    baseline; the fixed term is the span that differs from ``bottom_prompt``.
    Pairs whose swap span cannot be derived cleanly are skipped and recorded,
    never silently dropped. All baselines share one spend ceiling.
    """
    model = _resolve_model(model)
    client = _require_client(client)
    items: list[dict[str, Any]] = []
    baselines: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    all_variants: list[dict[str, str]] = []
    all_rejected: list[dict[str, Any]] = []
    rounds = 0
    spent = 0.0
    per_model: dict[str, dict[str, Any]] = {}
    truncated = False

    for pair in seed_pairs:
        if len(items) >= num_baselines:
            break
        phrase = str(pair.get("top_prompt") or "").strip()
        other = str(pair.get("bottom_prompt") or "").strip()
        term = _swap_span(phrase, other) if phrase and other else ""
        if not phrase or not term or term == phrase or len(term) < 3 or term not in phrase:
            skipped.append({"top_prompt": phrase, "reason": "no clean swap span"})
            continue
        remaining = max_spend - spent
        if remaining <= 0.01:
            truncated = True
            break
        result = generate_dialect_variants(
            phrase, term, n_variants,
            dialects=dialects, model=model, held_fixed="clinical",
            max_spend=remaining, client=client,
        )
        rounds += result["rounds"]
        spent += result["usage"]["total_cost_usd"]
        truncated = truncated or result["truncated"]
        for name, usage in result["usage"]["per_model"].items():
            slot = per_model.setdefault(name, {"calls": 0, "input_tokens": 0, "output_tokens": 0, "cost": 0.0})
            for key in slot:
                slot[key] = round(slot[key] + usage.get(key, 0), 6)
        all_rejected.extend(result["rejected"])
        if not result["variants"]:
            skipped.append({"top_prompt": phrase, "term": term, "reason": "no variants accepted"})
            continue
        all_variants.extend(result["variants"])
        target = pair.get("target_clinical_token")
        items.append(dialect_batch_item(result, target_token=target))
        baselines.append({
            "baseline_prompt": phrase,
            "term": term,
            "target_token": target,
            "variants": len(result["variants"]),
            "cost_usd": result["usage"]["total_cost_usd"],
        })

    return {
        "items": items,
        "baselines": baselines,
        "skipped": skipped,
        "variants": all_variants,
        "rejected": all_rejected,
        "model": model,
        "rounds": rounds,
        "truncated": truncated,
        "usage": {"total_cost_usd": round(spent, 6), "per_model": per_model},
    }


# ---------------------------------------------------------------------------
# 4. Spreadsheet importer
# ---------------------------------------------------------------------------

# Column layout of the hand-built sheet:
# Phrase 1 | Next token | Prob | Link | Phrase 2 | (next token) | (prob) | Link | Notes | sourcing
PHRASE1_COL, TOKEN1_COL, PROB1_COL = 0, 1, 2
PHRASE2_COL, TOKEN2_COL, PROB2_COL = 4, 5, 6
NOTES_COL, SOURCING_COL = 8, 9
SHEET_WIDTH = 10

LINK_RE = re.compile(r"https?://(?:www\.)?neuronpedia\.org/\S+")
# Trailing metadata row: first cell names the traced model / transcoder set.
METADATA_ROW_RE = re.compile(r"(?i)\b(gemma|qwen|llama|gpt|transcoder|gemmascope)\b")


def _read_rows(path: str | Path) -> Iterator[list[Any]]:
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError as e:
            raise RuntimeError(
                'reading .xlsx requires openpyxl - install with pip install "medlang-circuits[sheets]" '
                "(or export the sheet to CSV, which needs no extra dependency)"
            ) from e
        workbook = openpyxl.load_workbook(p, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        for row in worksheet.iter_rows(values_only=True):
            yield list(row)
    else:
        with open(p, newline="", encoding="utf-8-sig") as f:
            yield from csv.reader(f)


def _text(value: Any) -> str | None:
    """Outer-stripped cell text (interior preserved verbatim), or None if empty."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _prob(value: Any) -> float | str | None:
    """Parse a probability cell; unparseable non-empty values pass through as text."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return str(value).strip()


def _import_row(cells: list[Any], row_number: int) -> tuple[dict[str, Any] | None, str | None]:
    """Convert one sheet row to a batch pair; returns (pair, None) or (None, skip_reason)."""
    cells = (list(cells) + [None] * SHEET_WIDTH)[:SHEET_WIDTH]
    text = [_text(c) for c in cells]
    if not any(text):
        return None, "empty"

    # A neuronpedia.org URL is a circuit link wherever it appears (one row has
    # it parked in the Next-token column) - and never doubles as token/phrase text.
    links: dict[int, str] = {}
    for i, cell_text in enumerate(text):
        match = LINK_RE.search(cell_text or "")
        if match:
            links[i] = match.group(0)
            text[i] = None

    phrase1, phrase2 = text[PHRASE1_COL], text[PHRASE2_COL]
    if phrase1 and not phrase2 and METADATA_ROW_RE.search(phrase1):
        return None, f"metadata row ({phrase1!r})"
    if not phrase1 or not phrase2:
        return None, "missing one of the two phrases"

    patient_token, clinical_token = text[TOKEN1_COL], text[TOKEN2_COL]
    pair: dict[str, Any] = {
        "top_prompt": phrase2,  # Phrase 2 = clinical wording
        "bottom_prompt": phrase1,  # Phrase 1 = patient wording
    }
    if clinical_token:
        pair["target_clinical_token"] = " " + clinical_token
    # missing clinical next token -> unanchored pair: tracing falls back to top logits
    pair["provenance"] = {
        "source_row": row_number,
        "patient": {
            "observed_next_token": patient_token,
            "observed_prob": _prob(cells[PROB1_COL]),
            "circuit_link": next((links[i] for i in sorted(links) if i < PHRASE2_COL), None),
        },
        "clinical": {
            "observed_next_token": clinical_token,
            "observed_prob": _prob(cells[PROB2_COL]),
            "circuit_link": next((links[i] for i in sorted(links) if i >= PHRASE2_COL), None),
        },
        "notes": text[NOTES_COL],
        "sourcing": text[SOURCING_COL],
    }
    return pair, None


def import_sheet(path: str | Path) -> list[dict[str, Any]]:
    """Convert the hand-built spreadsheet (.xlsx via the 'sheets' extra, or CSV) to batch pairs.

    Forgiving by design: tokens are whitespace-stripped, fully-empty rows and
    the trailing model/transcoder metadata row are skipped, links are detected
    in any column, probs/links/tokens may be missing, and phrase text is kept
    verbatim (intentional misspellings included - they are part of the stress
    test). Manual measurements ride along under each pair's ``provenance``.
    """
    pairs: list[dict[str, Any]] = []
    for row_number, cells in enumerate(_read_rows(path), start=1):
        text = [_text(c) for c in (list(cells) + [None])[:1]]
        if row_number == 1 and text[0] and text[0].casefold().startswith("phrase"):
            continue  # header row
        pair, reason = _import_row(list(cells), row_number)
        if pair is None:
            if reason != "empty":
                logger.info("Skipping row %d: %s", row_number, reason)
            continue
        pairs.append(pair)
    logger.info("Imported %d pairs from %s", len(pairs), path)
    return pairs


# ---------------------------------------------------------------------------
# CLI: medlang-generate {pairs, dialects, import-sheet}
# ---------------------------------------------------------------------------


def _load_seed_pairs(path: str | None) -> list[dict[str, Any]] | None:
    if not path:
        return None
    with open(path, encoding="utf-8") as f:
        seeds = json.load(f)
    if not isinstance(seeds, list):
        raise ValueError(f"{path} must contain a JSON array of pair objects")
    return seeds


def _write_json(path: str | Path, payload: Any) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


def write_generation_report(out_path: str, task: str, max_spend: float,
                            result: dict[str, Any], extra: dict[str, Any] | None = None) -> Path:
    """Write the cost/provenance sidecar next to a generated batch.

    ``pairs_<stamp>.json`` gets ``pairs_<stamp>.report.json`` carrying the run
    timestamp, model, spend (actual vs. ceiling), accept/reject counts with
    reasons, and the run's steering parameters - so every archived batch is
    accountable without digging through CI logs.
    """
    accepted = result.get("pairs", result.get("variants", []))
    report = {
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "task": task,
        "model": result["model"],
        "accepted": len(accepted),
        "rejected": len(result["rejected"]),
        "rejection_reasons": sorted({r["reason"] for r in result["rejected"]}),
        "rounds": result["rounds"],
        "truncated_by_budget": result["truncated"],
        "max_spend_usd": max_spend,
        "cost_usd": result["usage"]["total_cost_usd"],
        "usage": result["usage"],
        "batch_file": out_path,
        **(extra or {}),
    }
    path = Path(out_path).with_suffix(".report.json")
    _write_json(path, report)
    return path


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    parser = argparse.ArgumentParser(
        prog="medlang-generate",
        description="Generate stress pairs / dialect variants with Claude, or import the hand-built sheet.",
    )
    shared = argparse.ArgumentParser(add_help=False)
    shared.add_argument("--model", default=None,
                        help="Anthropic model (default: MEDLANG_ANTHROPIC_MODEL or claude-opus-4-8)")
    shared.add_argument("--max-spend", type=float, default=2.0, help="Hard USD ceiling for the run")
    shared.add_argument("--out", default=None, help="Output JSON path (batch-ready pairs)")
    shared.add_argument("--seed-pairs", default=None,
                        help="JSON file of existing pairs (few-shot examples + dedupe set)")
    shared.add_argument("--feedback", default=None,
                        help="(pairs) JSON file of measurement-screening failures from a previous "
                             "round, shown to the generator as counterexamples")

    sub = parser.add_subparsers(dest="command", required=True)

    p_pairs = sub.add_parser("pairs", parents=[shared],
                             help="Generate validated patient-vs-clinical stress pairs")
    p_pairs.add_argument("-n", "--num", type=int, default=10, help="Pairs to accept")
    p_pairs.add_argument("--topics", nargs="+", default=None, help="Optional topics to steer coverage")

    p_quadrants = sub.add_parser("quadrants", parents=[shared],
                                 help="Generate validated 4-quadrant items: standard/nonstandard "
                                      "morphosyntax frames x medical/patient terms (--mode 4quadrant input)")
    p_quadrants.add_argument("-n", "--num", type=int, default=6, help="Items to accept")
    p_quadrants.add_argument("--topics", nargs="+", default=None, help="Optional topics to steer coverage")

    p_dialects = sub.add_parser("dialects", parents=[shared],
                                help="Generate dialect/register variants of one phrase (--mode dialect input)")
    p_dialects.add_argument("--phrase", default=None, help="Baseline phrase (the standard phrasing)")
    p_dialects.add_argument("--term", default=None, help="Swapped term to hold verbatim and fixed")
    p_dialects.add_argument("--num-baselines", type=int, default=8,
                            help="With --seed-pairs: how many measured baselines to sweep")
    p_dialects.add_argument("-n", "--num", type=int, default=6, help="Variants to accept")
    p_dialects.add_argument("--dialects", nargs="+", default=None,
                            help="Dialects/registers to draw from (default: built-in coverage list)")
    p_dialects.add_argument("--held-fixed", choices=["clinical", "patient"], default="clinical",
                            help="Which side of the swap the fixed term is (clinical isolates pure syntax)")
    p_dialects.add_argument("--target-token", default=None,
                            help="target_clinical_token for the emitted batch item")

    p_import = sub.add_parser("import-sheet", parents=[shared],
                              help="Convert the hand-built spreadsheet (.xlsx/.csv) to batch pairs")
    p_import.add_argument("sheet", help="Path to the .xlsx (needs the 'sheets' extra) or .csv file")

    args = parser.parse_args(sys.argv[1:] if argv is None else argv)

    if args.command == "import-sheet":
        pairs = import_sheet(args.sheet)
        out = args.out or "imported_pairs.json"
        _write_json(out, pairs)
        print(json.dumps({"imported": len(pairs), "out": out}, indent=2))
        return 0

    if args.command == "pairs":
        feedback = None
        if args.feedback:
            with open(args.feedback, encoding="utf-8") as f:
                feedback = json.load(f)
        result = generate_stress_pairs(
            args.num,
            model=args.model,
            seed_pairs=_load_seed_pairs(args.seed_pairs),
            topics=args.topics,
            max_spend=args.max_spend,
            feedback=feedback,
        )
        out = args.out or "generated_pairs.json"
        _write_json(out, result["pairs"])
        # dialect-tagged registers are a later capability; today's pairs are
        # general patient language, and the report says so explicitly
        extra: dict[str, Any] = {"language_register": "general_patient_language",
                                 "topics": result["topics"]}
    elif args.command == "quadrants":
        result = generate_quadrant_scenarios(
            args.num,
            model=args.model,
            seed_pairs=_load_seed_pairs(args.seed_pairs),
            topics=args.topics,
            max_spend=args.max_spend,
        )
        out = args.out or "quadrant_scenarios.json"
        _write_json(out, result["pairs"])
        extra = {"language_register": "morphosyntax_x_lexicon_quadrants",
                 "topics": result["topics"]}
    elif args.command == "dialects" and args.seed_pairs:
        result = generate_dialect_sweep(
            _load_seed_pairs(args.seed_pairs) or [],
            args.num_baselines,
            args.num,
            dialects=args.dialects,
            model=args.model,
            max_spend=args.max_spend,
        )
        out = args.out or "dialect_pairs.json"
        _write_json(out, result["items"])
        extra = {"held_fixed": "clinical", "baselines": result["baselines"],
                 "skipped": result["skipped"]}
    else:  # dialects, single baseline
        if not args.phrase or not args.term:
            parser.error("dialects needs --phrase and --term, or --seed-pairs for a sweep")
        result = generate_dialect_variants(
            args.phrase,
            args.term,
            args.num,
            dialects=args.dialects,
            model=args.model,
            held_fixed=args.held_fixed,
            max_spend=args.max_spend,
        )
        out = args.out or "dialect_pairs.json"
        _write_json(out, [dialect_batch_item(result, target_token=args.target_token)])
        extra = {"baseline_prompt": result["baseline_prompt"], "term": result["term"],
                 "held_fixed": result["held_fixed"],
                 "dialects": [v["dialect"] for v in result["variants"]]}

    report_path = write_generation_report(out, args.command, args.max_spend, result, extra)
    print(json.dumps({
        "accepted": len(result.get("pairs", result.get("variants", []))),
        "rejected": len(result["rejected"]),
        "rounds": result["rounds"],
        "truncated_by_budget": result["truncated"],
        "cost_usd": result["usage"]["total_cost_usd"],
        "max_spend_usd": args.max_spend,
        "usage": result["usage"],
        "out": out,
        "report": str(report_path),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
